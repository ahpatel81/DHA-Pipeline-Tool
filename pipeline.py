"""
Verve Pipeline Intelligence Tool
=================================
Extracts contracts from an OrangeSlices AI article and enriches each one
with incumbent details, spending data, SOW links, and related news using
the Claude AI API + web search.

Usage:
    python pipeline.py --url "https://orangeslices.ai/..." --agency DHA
    python pipeline.py --url "https://orangeslices.ai/..." --agency VA

Output:
    An Excel file in ./output/ named by agency and retrieval date.
    Example: output/VA_2026-04-02.xlsx

Requirements:
    pip install anthropic openpyxl requests beautifulsoup4

You need an Anthropic API key set as an environment variable:
    export ANTHROPIC_API_KEY="sk-ant-..."
"""

import anthropic
import json
import re
import sys
import os
import argparse
from datetime import date
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── CONFIG ────────────────────────────────────────────────────────────────────

MODEL = "claude-opus-4-5"          # Use the most capable model for research
MAX_TOKENS = 1500
OUTPUT_DIR = Path("output")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


# ── STEP 1: FETCH & PARSE ARTICLE ────────────────────────────────────────────

def fetch_article(url: str) -> str:
    """Download the OrangeSlices article and return its plain text."""
    print(f"  Fetching article: {url}")
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    # Remove nav/footer noise
    for tag in soup(["nav", "footer", "script", "style", "header"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True)


def extract_contracts(article_text: str, agency: str, client: anthropic.Anthropic) -> list[dict]:
    """
    Use Claude to extract all contracts from the article text.
    Returns a list of dicts with: title, contract_id, value_m, bidders
    """
    print("  Extracting contracts from article...")
    prompt = f"""
Extract every contract or task order listed in this article. The article is about
{agency} expiring contracts. Each entry typically looks like:
  $123.4M SOME CONTRACT TITLE | CONTRACT-ID-HERE | Number of Bidders: 5

Return ONLY a JSON array. No markdown, no preamble. Each object must have:
  - "title": string (the contract title, ALL CAPS is fine)
  - "contract_id": string (the contract/task order number)
  - "value_m": number (dollar value in millions, e.g. 123.4)
  - "bidders": number (integer)

Article text:
---
{article_text[:12000]}
---
"""
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = response.content[0].text.strip()
    raw = re.sub(r"^```json|^```|```$", "", raw, flags=re.MULTILINE).strip()
    contracts = json.loads(raw)
    print(f"  Found {len(contracts)} contracts.")
    return contracts


# ── STEP 2: ENRICH EACH CONTRACT ─────────────────────────────────────────────

ENRICH_PROMPT = """
You are a federal contract intelligence analyst. Research this {agency} contract and
return structured intelligence. Use your knowledge of federal contracting, USASpending,
SAM.gov, and GovCon news to fill in as much as possible.

Contract: {title}
Contract ID: {contract_id}
Value: ${value_m}M
Number of bidders at last award: {bidders}
Agency: {agency}

Return ONLY a JSON object (no markdown, no extra text) with these exact keys:

{{
  "incumbent": "Name of current awardee, or 'Unknown'",
  "incumbent_notes": "1-2 sentences on their past performance or relevant background",
  "contract_type": "e.g. Task Order, IDIQ, BPA, FFP, Cost-Plus",
  "naics_code": "Most likely NAICS code",
  "naics_desc": "NAICS description",
  "set_aside": "e.g. Full and Open, 8(a), SDVOSB, HUBZone, Small Business",
  "usaspending_note": "What USASpending.gov would show — obligated vs ceiling, mods, etc.",
  "sam_note": "What to expect on SAM.gov — solicitation status, last posted date, etc.",
  "recompete_risk": "Low / Medium / High",
  "recompete_rationale": "1 sentence explaining the risk level",
  "related_news": "Any known press releases, budget exhibits, or news referencing this program",
  "capture_actions": "Top 2-3 concrete capture actions for a BD team"
}}
"""

def enrich_contract(contract: dict, agency: str, client: anthropic.Anthropic) -> dict:
    """Call Claude with web search to enrich a single contract."""
    print(f"    Enriching: {contract['title'][:60]}...")
    prompt = ENRICH_PROMPT.format(
        agency=agency,
        title=contract["title"],
        contract_id=contract["contract_id"],
        value_m=contract["value_m"],
        bidders=contract["bidders"],
    )
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        tools=[{"type": "web_search_20250305", "name": "web_search"}],
        messages=[{"role": "user", "content": prompt}]
    )
    # Collect text blocks (model may call web_search then respond)
    text = ""
    for block in response.content:
        if block.type == "text":
            text += block.text
    text = text.strip()
    text = re.sub(r"^```json|^```|```$", "", text, flags=re.MULTILINE).strip()
    try:
        enriched = json.loads(text)
    except json.JSONDecodeError:
        enriched = {
            "incumbent": "Parse error",
            "incumbent_notes": text[:200],
            "contract_type": "", "naics_code": "", "naics_desc": "",
            "set_aside": "", "usaspending_note": "", "sam_note": "",
            "recompete_risk": "", "recompete_rationale": "",
            "related_news": "", "capture_actions": ""
        }
    return {**contract, **enriched}


# ── STEP 3: WRITE EXCEL OUTPUT ────────────────────────────────────────────────

COLUMNS = [
    ("Contract Title",         40),
    ("Contract ID",            28),
    ("Value ($M)",             10),
    ("Bidders",                 8),
    ("Agency",                 12),
    ("Date Retrieved",         14),
    ("Incumbent",              24),
    ("Incumbent Notes",        40),
    ("Contract Type",          16),
    ("NAICS Code",             12),
    ("NAICS Description",      28),
    ("Set-Aside",              16),
    ("USASpending Note",       44),
    ("SAM.gov Note",           36),
    ("Recompete Risk",         14),
    ("Recompete Rationale",    40),
    ("Related News",           44),
    ("Capture Actions",        44),
    ("HigherGov Link",         48),
    ("USASpending Link",       48),
]

HDR_FILL   = PatternFill("solid", fgColor="1F3864")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=10)
ALT_FILL   = PatternFill("solid", fgColor="EEF3FA")
WRAP       = Alignment(wrap_text=True, vertical="top")
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RISK_COLORS = {"High": "FCE4D6", "Medium": "FFF2CC", "Low": "E2EFDA"}

def highergov_url(contract_id: str) -> str:
    return f"https://www.highergov.com/contract/{contract_id}/?ref=os"

def usaspending_url(contract_id: str) -> str:
    return f"https://www.usaspending.gov/search/?query={contract_id}"

def write_excel(rows: list[dict], agency: str, retrieved: str, output_path: Path):
    wb = openpyxl.Workbook()

    # ── Master sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = f"{agency}_{retrieved}"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = HDR_FONT
        cell.fill    = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border  = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        alt = (row_idx % 2 == 0)
        risk = row.get("recompete_risk", "")
        risk_fill = PatternFill("solid", fgColor=RISK_COLORS.get(risk, "FFFFFF")) if risk else None

        values = [
            row.get("title", ""),
            row.get("contract_id", ""),
            row.get("value_m", ""),
            row.get("bidders", ""),
            agency,
            retrieved,
            row.get("incumbent", ""),
            row.get("incumbent_notes", ""),
            row.get("contract_type", ""),
            row.get("naics_code", ""),
            row.get("naics_desc", ""),
            row.get("set_aside", ""),
            row.get("usaspending_note", ""),
            row.get("sam_note", ""),
            risk,
            row.get("recompete_rationale", ""),
            row.get("related_news", ""),
            row.get("capture_actions", ""),
            highergov_url(row.get("contract_id", "")),
            usaspending_url(row.get("contract_id", "")),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = BODY_FONT
            cell.alignment = WRAP
            cell.border    = BORDER
            # Recompete Risk column gets color
            if col_idx == 15 and risk_fill:
                cell.fill = risk_fill
            elif alt:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 60

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ss = wb.create_sheet("Summary")
    ss.column_dimensions["A"].width = 32
    ss.column_dimensions["B"].width = 24

    summary_rows = [
        ("Agency",               agency),
        ("Date Retrieved",       retrieved),
        ("Total Contracts",      f'=COUNTA(\'{ws.title}\'!A2:A1000)'),
        ("Total Value ($M)",     f'=SUM(\'{ws.title}\'!C2:C1000)'),
        ("Avg Bidders",          f'=AVERAGE(\'{ws.title}\'!D2:D1000)'),
        ("High Recompete Risk",  f'=COUNTIF(\'{ws.title}\'!O2:O1000,"High")'),
        ("Medium Recompete Risk",f'=COUNTIF(\'{ws.title}\'!O2:O1000,"Medium")'),
        ("Low Recompete Risk",   f'=COUNTIF(\'{ws.title}\'!O2:O1000,"Low")'),
    ]

    for r, (label, value) in enumerate(summary_rows, 2):
        lc = ss.cell(row=r, column=1, value=label)
        vc = ss.cell(row=r, column=2, value=value)
        lc.font = Font(name="Arial", bold=True, size=10)
        vc.font = Font(name="Arial", size=10)
        lc.border = vc.border = BORDER

    hdr = ss.cell(row=1, column=1, value="Pipeline Summary")
    hdr.font = Font(name="Arial", bold=True, size=13, color="1F3864")
    ss.merge_cells("A1:B1")

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Verve Pipeline Intelligence Tool")
    parser.add_argument("--url",    required=True,  help="OrangeSlices article URL")
    parser.add_argument("--agency", required=True,  help="Agency name, e.g. VA or DHA")
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: Set ANTHROPIC_API_KEY environment variable first.")
        print("  export ANTHROPIC_API_KEY='sk-ant-...'")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    retrieved = date.today().isoformat()   # e.g. 2026-04-02
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / f"{args.agency}_{retrieved}.xlsx"

    print(f"\n=== Verve Pipeline Tool ===")
    print(f"Agency:    {args.agency}")
    print(f"Retrieved: {retrieved}")
    print(f"Output:    {output_path}\n")

    print("[1/3] Fetching article...")
    article_text = fetch_article(args.url)

    print("[2/3] Extracting contracts...")
    contracts = extract_contracts(article_text, args.agency, client)

    print("[3/3] Enriching contracts (this takes a few minutes)...")
    enriched_rows = []
    for i, contract in enumerate(contracts, 1):
        print(f"  [{i}/{len(contracts)}]", end=" ")
        enriched = enrich_contract(contract, args.agency, client)
        enriched_rows.append(enriched)

    print("\n[4/4] Writing Excel output...")
    write_excel(enriched_rows, args.agency, retrieved, output_path)

    print("\nDone. Open the file in Excel or Google Sheets.")
    print(f"To run again on a new list: python pipeline.py --url <URL> --agency <AGENCY>")


if __name__ == "__main__":
    main()
